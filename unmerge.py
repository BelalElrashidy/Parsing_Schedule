from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
def unmerge(file:str) -> str:
    wb = load_workbook(filename = file)
    new_file = file +"_modified.xlsx"
    for st_name in wb.sheetnames:
        st = wb[st_name]
        mcr_coord_list = [mcr.coord for mcr in st.merged_cells.ranges]
        
        for mcr in mcr_coord_list:
            min_col, min_row, max_col, max_row = range_boundaries(mcr)
            top_left_cell_value = st.cell(row=min_row, column=min_col).value
            st.unmerge_cells(mcr)
            for row in st.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value

    wb.save(new_file)
    return new_file